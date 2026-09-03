"""
Servicio de intercambio y portabilidad de datos para CAIT Informes.
Permite exportar e importar datos clínicos, pacientes, informes completos y bases de datos
en formatos no-PDF (JSON/.cait, Excel/.xlsx, CSV/.csv, y paquetes .caitbackup/.caitpkg),
asegurando el auto-registro de la información en la base de datos de la máquina destino.
"""

from __future__ import annotations

import base64
import csv
import io
import json
import os
import re
import shutil
import tempfile
import unicodedata
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class DataExchangeService:
    """Maneja la exportación e importación de datos clínicos entre aplicaciones y equipos."""

    VERSION = "2.3.0"
    FORMAT_IDENTIFIER = "CAIT_DATA_PACKAGE_V1"

    # =========================================================================
    # EXPORTACIÓN
    # =========================================================================

    def _collect_all_report_files(self, report_data: dict, data_root: Path) -> List[Dict[str, Any]]:
        """
        Recolecta todos los archivos físicos (PDFs, imágenes, certificados, idoneidades)
        asociados a un reporte para empaquetarlos en .caitpkg o embeberlos en .cait.
        """
        collected: List[Dict[str, Any]] = []
        seen_paths = set()

        def add_file(src_path: Path, category: str = "report_adjuntos", tipo: str = ""):
            if not src_path or not src_path.exists() or not src_path.is_file():
                return
            resolved = str(src_path.resolve())
            if resolved in seen_paths:
                return
            seen_paths.add(resolved)
            
            subfolder = category
            try:
                if src_path.resolve().is_relative_to((data_root / "attachments").resolve()):
                    rel_p = src_path.resolve().relative_to((data_root / "attachments").resolve())
                    sub = str(rel_p.parent).replace("\\", "/")
                    if sub and sub != ".":
                        subfolder = sub
            except Exception:
                pass

            collected.append({
                "name": src_path.name,
                "path": src_path,
                "category": subfolder,
                "tipo": tipo,
                "arcname": f"attachments/{subfolder}/{src_path.name}",
                "rel_path": f"attachments/{subfolder}/{src_path.name}"
            })

        # 1. Adjuntos de la lista 'adjuntos'
        for a in report_data.get("adjuntos", []):
            if not isinstance(a, dict):
                continue
            name = a.get("name")
            tipo = a.get("tipo", "")
            path_str = a.get("path")
            
            found = False
            if path_str:
                for candidate in [
                    data_root / path_str,
                    data_root / path_str.replace("data/", "").replace("data\\", ""),
                    Path(path_str)
                ]:
                    if candidate.exists() and candidate.is_file():
                        add_file(candidate, category="report_adjuntos", tipo=tipo)
                        found = True
                        break
            
            if not found and name:
                for folder in ["report_adjuntos", "idoneidad", "demo_zip_certificados", "demo_zip_resultados", "demo_asistencia", "demo_calibracion", "demo_protocolo", "demo_resultados"]:
                    candidate = data_root / "attachments" / folder / name
                    if candidate.exists() and candidate.is_file():
                        add_file(candidate, category=folder, tipo=tipo)
                        found = True
                        break

                if not found:
                    for f in (data_root / "attachments").rglob(name):
                        if f.is_file():
                            add_file(f, category="report_adjuntos", tipo=tipo)
                            break

        # 2. calibration_files
        for cf in report_data.get("calibration_files", []):
            p_str = cf if isinstance(cf, str) else (cf.get("file") or cf.get("name") if isinstance(cf, dict) else "")
            if p_str:
                for candidate in [data_root / "attachments" / "calibracion" / os.path.basename(p_str), data_root / p_str, Path(p_str)]:
                    if candidate.exists() and candidate.is_file():
                        add_file(candidate, category="calibracion", tipo="Certificado Calibración")
                        break

        # 3. attendance_files
        for af in report_data.get("attendance_files", []):
            p_str = af if isinstance(af, str) else (af.get("file") or af.get("name") if isinstance(af, dict) else "")
            if p_str:
                for candidate in [data_root / "attachments" / "asistencia" / os.path.basename(p_str), data_root / p_str, Path(p_str)]:
                    if candidate.exists() and candidate.is_file():
                        add_file(candidate, category="asistencia", tipo="Listado Asistencia")
                        break

        # 4. result_attachments
        res_att = report_data.get("result_attachments", {})
        if isinstance(res_att, dict):
            for aud in res_att.get("audiometria", []):
                p_str = aud if isinstance(aud, str) else (aud.get("file") if isinstance(aud, dict) else "")
                if p_str:
                    c = data_root / p_str if (data_root / p_str).exists() else (data_root / "attachments" / "report_adjuntos" / os.path.basename(p_str))
                    add_file(c, category="report_adjuntos", tipo="Audiograma")
            for esp in res_att.get("espirometria", []):
                p_str = esp if isinstance(esp, str) else (esp.get("file") if isinstance(esp, dict) else "")
                if p_str:
                    c = data_root / p_str if (data_root / p_str).exists() else (data_root / "attachments" / "report_adjuntos" / os.path.basename(p_str))
                    add_file(c, category="report_adjuntos", tipo="Reporte Espirometría")

        # 5. Idoneidades de evaluadores
        eval_dir = data_root / "attachments" / "idoneidad"
        if eval_dir.exists():
            for f in eval_dir.glob("*"):
                if f.is_file():
                    add_file(f, category="idoneidad", tipo="Idoneidad Profesional")

        return collected

    def _collect_all_drafts(self, data_root: Path) -> List[Dict[str, Any]]:
        """Recolecta todos los borradores almacenados en data_root/reports."""
        drafts: List[Dict[str, Any]] = []
        reports_dir = data_root / "reports"
        if reports_dir.exists():
            for f in sorted(reports_dir.glob("*.json")):
                try:
                    content = json.loads(f.read_text(encoding="utf-8"))
                    drafts.append({
                        "name": f.name,
                        "modified": os.path.getmtime(f),
                        "content": content
                    })
                except Exception as e:
                    print(f"Error leyendo borrador {f.name}: {e}")
        return drafts

    def export_report_cait(
        self,
        report_data: dict,
        data_root: Optional[Path] = None,
        persons_repo=None,
        include_files: bool = True,
        include_drafts: bool = True
    ) -> dict:
        """
        Genera una estructura serializable completa (.cait) del informe actual,
        incluyendo pacientes asociados, borradores y archivos/PDFs embebidos en base64
        para facilitar el auto-registro y portabilidad completa en otra PC.
        """
        data = dict(report_data)
        company = str(data.get("company_name") or data.get("company") or "Empresa").strip()
        eval_date = str(data.get("evaluation_date") or data.get("study_date") or datetime.now().strftime("%Y-%m-%d"))
        
        # Recolectar pacientes asociados para incluirlos en el paquete
        associated_persons = []
        if persons_repo:
            cedulas = set()
            for item in data.get("resultados_audiometria", []):
                if isinstance(item, dict) and item.get("cedula"):
                    cedulas.add(item["cedula"].strip().upper())
            for item in data.get("resultados_espirometria", []):
                if isinstance(item, dict) and item.get("cedula"):
                    cedulas.add(item["cedula"].strip().upper())
            
            for ced in cedulas:
                p = persons_repo.get_by_id(ced)
                if p:
                    associated_persons.append(p)

        # Archivos y PDFs embebidos
        embedded_files = []
        if include_files and data_root and data_root.exists():
            files_to_embed = self._collect_all_report_files(data, data_root)
            for finfo in files_to_embed:
                fpath = finfo["path"]
                try:
                    raw_bytes = fpath.read_bytes()
                    if len(raw_bytes) <= 20 * 1024 * 1024:
                        embedded_files.append({
                            "name": finfo["name"],
                            "category": finfo["category"],
                            "tipo": finfo.get("tipo", ""),
                            "rel_path": finfo["rel_path"],
                            "b64": base64.b64encode(raw_bytes).decode("ascii"),
                            "size": len(raw_bytes)
                        })
                except Exception as e:
                    print(f"Error embebiendo archivo {fpath}: {e}")

        # Borradores guardados en el sistema
        saved_drafts = []
        if include_drafts and data_root and data_root.exists():
            saved_drafts = self._collect_all_drafts(data_root)

        export_package = {
            "_header": {
                "format": self.FORMAT_IDENTIFIER,
                "version": self.VERSION,
                "exported_at": datetime.now().isoformat(),
                "app": "CAIT Informes Panamá",
            },
            "report": data,
            "associated_persons": associated_persons,
            "embedded_files": embedded_files,
            "saved_drafts": saved_drafts,
            "summary": {
                "company": company,
                "evaluation_date": eval_date,
                "report_type": data.get("report_type", "audiometria"),
                "total_audiometria": len(data.get("resultados_audiometria", [])),
                "total_espirometria": len(data.get("resultados_espirometria", [])),
                "total_files": len(embedded_files),
                "total_drafts": len(saved_drafts),
            }
        }
        return export_package

    def export_report_package_zip(
        self,
        report_data: dict,
        data_root: Path,
        persons_repo=None,
        include_drafts: bool = True
    ) -> Tuple[Path, str]:
        """
        Crea un paquete .caitpkg (ZIP portable) que incluye el JSON de datos,
        todos los archivos adjuntos y PDFs (certificados, idoneidades, etc.)
        y todos los borradores guardados para migrar todo sin perder nada.
        """
        cait_data = self.export_report_cait(
            report_data,
            data_root=data_root,
            persons_repo=persons_repo,
            include_files=False,  # En ZIP van como archivos binarios reales
            include_drafts=include_drafts
        )
        company = str(report_data.get("company_name") or "Informe").strip()
        nfkd = unicodedata.normalize('NFKD', company)
        ascii_comp = nfkd.encode('ascii', 'ignore').decode('ascii')
        safe_company = "".join(c for c in ascii_comp if c.isalnum() or c in (" ", "_", "-")).strip().replace(" ", "_") or "Informe"
        filename = f"CAIT_{safe_company}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.caitpkg"
        
        temp_dir = Path(tempfile.mkdtemp(prefix="cait_pkg_"))
        pkg_zip_path = temp_dir / filename
        
        with zipfile.ZipFile(pkg_zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            # Guardar el JSON principal
            zf.writestr("manifest.json", json.dumps(cait_data, ensure_ascii=False, indent=2))
            
            # Recolectar y guardar todos los adjuntos físicos (PDFs, imágenes, etc.)
            files_to_pack = self._collect_all_report_files(report_data, data_root)
            for finfo in files_to_pack:
                src = finfo["path"]
                if src.exists():
                    zf.write(src, arcname=finfo["arcname"])
            
            # Recolectar y guardar todos los borradores del sistema en carpeta drafts/
            if include_drafts:
                reports_dir = data_root / "reports"
                if reports_dir.exists():
                    for draft_file in reports_dir.glob("*.json"):
                        zf.write(draft_file, arcname=f"drafts/{draft_file.name}")
                        
            # Si existe informe PDF generado en exports, incluirlo también
            exports_dir = data_root / "exports"
            if exports_dir.exists():
                for exp_pdf in exports_dir.rglob("*.pdf"):
                    if exp_pdf.is_file() and safe_company.lower() in exp_pdf.name.lower():
                        zf.write(exp_pdf, arcname=f"exports/{exp_pdf.name}")
                        break

        return pkg_zip_path, filename

    def export_to_excel(self, report_data: dict, persons_repo=None) -> io.BytesIO:
        """
        Exporta los datos completos del informe a un libro de Excel (.xlsx)
        con formato profesional, cabeceras estilizadas y múltiples hojas.
        """
        wb = openpyxl.Workbook()
        # Eliminar hoja por defecto
        default_sheet = wb.active
        
        # Paleta de colores CAIT
        header_fill = PatternFill(start_color="00450D", end_color="00450D", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
        title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        sub_fill = PatternFill(start_color="EAE7E7", end_color="EAE7E7", fill_type="solid")
        sub_font = Font(name="Calibri", size=10, bold=True, color="1B1C1C")
        regular_font = Font(name="Calibri", size=10)
        bold_font = Font(name="Calibri", size=10, bold=True)
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC")
        )

        # ---------------------------------------------------------------------
        # HOJA 1: RESUMEN GENERAL DEL INFORME
        # ---------------------------------------------------------------------
        ws_info = wb.create_sheet(title="Resumen del Informe")
        ws_info.views.sheetView[0].showGridLines = True
        
        ws_info.merge_cells("A1:D1")
        top_cell = ws_info["A1"]
        top_cell.value = "CAIT PANAMÁ — INFORME CLÍNICO OCUPACIONAL"
        top_cell.fill = title_fill
        top_cell.font = title_font
        top_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_info.row_dimensions[1].height = 32

        # Información general
        info_rows = [
            ("Empresa:", report_data.get("company_name", "N/A"), "Tipo de Informe:", report_data.get("report_type", "Audiometría")),
            ("Fecha Evaluación:", report_data.get("evaluation_date") or report_data.get("evaluation_dates", "N/A"), "Fecha del Estudio:", report_data.get("study_date") or report_data.get("study_dates", "N/A")),
            ("Ubicación / Área:", report_data.get("location", "N/A"), "Actividad Principal:", report_data.get("company_activity", "N/A")),
            ("País:", report_data.get("country", "Panamá"), "Dirección:", report_data.get("company_address", "N/A")),
            ("Contraparte Técnica:", report_data.get("counterpart_name", "N/A"), "Cargo Contraparte:", report_data.get("counterpart_role", "N/A")),
            ("Evaluador Principal:", report_data.get("evaluator_main", "N/A"), "Exportado el:", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ]

        curr_row = 3
        ws_info.cell(row=curr_row, column=1, value="DATOS GENERALES DEL INFORME").font = Font(name="Calibri", size=11, bold=True, color="00450D")
        ws_info.merge_cells(f"A{curr_row}:D{curr_row}")
        curr_row += 1

        for label1, val1, label2, val2 in info_rows:
            c1 = ws_info.cell(row=curr_row, column=1, value=label1)
            c1.font = bold_font
            c1.fill = sub_fill
            c1.border = thin_border
            
            c2 = ws_info.cell(row=curr_row, column=2, value=str(val1 or ""))
            c2.font = regular_font
            c2.border = thin_border
            
            c3 = ws_info.cell(row=curr_row, column=3, value=label2)
            c3.font = bold_font
            c3.fill = sub_fill
            c3.border = thin_border
            
            c4 = ws_info.cell(row=curr_row, column=4, value=str(val2 or ""))
            c4.font = regular_font
            c4.border = thin_border
            curr_row += 1

        curr_row += 1
        # Conclusiones y recomendaciones
        ws_info.cell(row=curr_row, column=1, value="CONCLUSIÓN").font = Font(name="Calibri", size=11, bold=True, color="00450D")
        ws_info.merge_cells(f"A{curr_row}:D{curr_row}")
        curr_row += 1
        
        c_conc = ws_info.cell(row=curr_row, column=1, value=report_data.get("conclusion_text") or report_data.get("conclusion", "Sin conclusión registrada."))
        c_conc.font = regular_font
        c_conc.alignment = Alignment(wrap_text=True, vertical="top")
        ws_info.merge_cells(f"A{curr_row}:D{curr_row + 3}")
        curr_row += 5

        ws_info.cell(row=curr_row, column=1, value="RECOMENDACIONES").font = Font(name="Calibri", size=11, bold=True, color="00450D")
        ws_info.merge_cells(f"A{curr_row}:D{curr_row}")
        curr_row += 1
        
        c_rec = ws_info.cell(row=curr_row, column=1, value=report_data.get("recommendations_text") or report_data.get("recommendations", "Sin recomendaciones registradas."))
        c_rec.font = regular_font
        c_rec.alignment = Alignment(wrap_text=True, vertical="top")
        ws_info.merge_cells(f"A{curr_row}:D{curr_row + 3}")

        # ---------------------------------------------------------------------
        # HOJA 2: AUDIOMETRÍA
        # ---------------------------------------------------------------------
        audio_rows = report_data.get("resultados_audiometria", [])
        ws_audio = wb.create_sheet(title="Audiometría")
        ws_audio.views.sheetView[0].showGridLines = True
        
        ws_audio.cell(row=1, column=1, value="RESULTADOS DE EVALUACIÓN AUDIOMÉTRICA").font = title_font
        ws_audio.cell(row=1, column=1).fill = title_fill
        ws_audio.merge_cells("A1:F1")
        ws_audio.row_dimensions[1].height = 28
        
        headers_audio = ["N°", "Nombre Completo", "Cédula", "Edad", "Área / Puesto", "Diagnóstico / Resultado"]
        for col_idx, h in enumerate(headers_audio, start=1):
            cell = ws_audio.cell(row=2, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center" if col_idx in (1, 3, 4) else "left", vertical="center")
            cell.border = thin_border
        ws_audio.row_dimensions[2].height = 24

        for r_idx, item in enumerate(audio_rows, start=1):
            row_num = r_idx + 2
            ws_audio.cell(row=row_num, column=1, value=r_idx).alignment = Alignment(horizontal="center")
            ws_audio.cell(row=row_num, column=2, value=item.get("name", ""))
            ws_audio.cell(row=row_num, column=3, value=item.get("cedula", "")).alignment = Alignment(horizontal="center")
            ws_audio.cell(row=row_num, column=4, value=item.get("age", "")).alignment = Alignment(horizontal="center")
            ws_audio.cell(row=row_num, column=5, value=item.get("position", ""))
            
            res_cell = ws_audio.cell(row=row_num, column=6, value=item.get("result", ""))
            # Color distintivo en resultado
            res_lower = str(item.get("result", "")).lower()
            if "normal" in res_lower:
                res_cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                res_cell.font = Font(name="Calibri", size=10, bold=True, color="1B5E20")
            elif any(k in res_lower for k in ("caída", "anacusia", "alterad", "sever", "profund")):
                res_cell.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                res_cell.font = Font(name="Calibri", size=10, bold=True, color="B71C1C")
            
            for c in range(1, 7):
                ws_audio.cell(row=row_num, column=c).border = thin_border

        # ---------------------------------------------------------------------
        # HOJA 3: ESPIROMETRÍA
        # ---------------------------------------------------------------------
        espiro_rows = report_data.get("resultados_espirometria", [])
        ws_espiro = wb.create_sheet(title="Espirometría")
        ws_espiro.views.sheetView[0].showGridLines = True
        
        ws_espiro.cell(row=1, column=1, value="RESULTADOS DE EVALUACIÓN ESPIROMÉTRICA").font = title_font
        ws_espiro.cell(row=1, column=1).fill = title_fill
        ws_espiro.merge_cells("A1:F1")
        ws_espiro.row_dimensions[1].height = 28
        
        headers_espiro = ["N°", "Nombre Completo", "Cédula", "Edad", "Área / Puesto", "Diagnóstico / Resultado"]
        for col_idx, h in enumerate(headers_espiro, start=1):
            cell = ws_espiro.cell(row=2, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center" if col_idx in (1, 3, 4) else "left", vertical="center")
            cell.border = thin_border
        ws_espiro.row_dimensions[2].height = 24

        for r_idx, item in enumerate(espiro_rows, start=1):
            row_num = r_idx + 2
            ws_espiro.cell(row=row_num, column=1, value=r_idx).alignment = Alignment(horizontal="center")
            ws_espiro.cell(row=row_num, column=2, value=item.get("name", ""))
            ws_espiro.cell(row=row_num, column=3, value=item.get("cedula", "")).alignment = Alignment(horizontal="center")
            ws_espiro.cell(row=row_num, column=4, value=item.get("age", "")).alignment = Alignment(horizontal="center")
            ws_espiro.cell(row=row_num, column=5, value=item.get("position", ""))
            
            res_cell = ws_espiro.cell(row=row_num, column=6, value=item.get("result", ""))
            res_lower = str(item.get("result", "")).lower()
            if "normal" in res_lower:
                res_cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                res_cell.font = Font(name="Calibri", size=10, bold=True, color="1B5E20")
            elif any(k in res_lower for k in ("restricción", "obstrucción", "alterad", "grave")):
                res_cell.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                res_cell.font = Font(name="Calibri", size=10, bold=True, color="B71C1C")
                
            for c in range(1, 7):
                ws_espiro.cell(row=row_num, column=c).border = thin_border

        # ---------------------------------------------------------------------
        # HOJA 4: CATÁLOGO DE PACIENTES / PERSONAS REGISTRADAS
        # ---------------------------------------------------------------------
        if persons_repo:
            all_persons = persons_repo.list_all()
            if all_persons:
                ws_persons = wb.create_sheet(title="Registro de Pacientes")
                ws_persons.views.sheetView[0].showGridLines = True
                
                ws_persons.cell(row=1, column=1, value="CATÁLOGO MAESTRO DE PERSONAS EVALUADAS").font = title_font
                ws_persons.cell(row=1, column=1).fill = title_fill
                ws_persons.merge_cells("A1:F1")
                ws_persons.row_dimensions[1].height = 28
                
                headers_p = ["N°", "Cédula", "Nombre Completo", "Edad", "Área / Puesto", "Último Resultado"]
                for col_idx, h in enumerate(headers_p, start=1):
                    cell = ws_persons.cell(row=2, column=col_idx, value=h)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center" if col_idx in (1, 2, 4) else "left", vertical="center")
                    cell.border = thin_border
                ws_persons.row_dimensions[2].height = 24
                
                for r_idx, p in enumerate(all_persons, start=1):
                    row_num = r_idx + 2
                    ws_persons.cell(row=row_num, column=1, value=r_idx).alignment = Alignment(horizontal="center")
                    ws_persons.cell(row=row_num, column=2, value=p.get("identification", "")).alignment = Alignment(horizontal="center")
                    ws_persons.cell(row=row_num, column=3, value=p.get("name", ""))
                    ws_persons.cell(row=row_num, column=4, value=p.get("age", "")).alignment = Alignment(horizontal="center")
                    ws_persons.cell(row=row_num, column=5, value=p.get("position", ""))
                    ws_persons.cell(row=row_num, column=6, value=p.get("last_result_label", ""))
                    for c in range(1, 7):
                        ws_persons.cell(row=row_num, column=c).border = thin_border

        # Eliminar hoja inicial vacía si existe
        if default_sheet in wb.worksheets:
            wb.remove(default_sheet)

        # Ajustar ancho de columnas automáticamente en todas las hojas
        for ws in wb.worksheets:
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    # Ignorar celdas fusionadas muy largas
                    if cell.coordinate in ("A1", "B1", "C1", "D1", "E1", "F1") and ws.merged_cells:
                        continue
                    val_str = str(cell.value or "")
                    if "\n" in val_str:
                        lines = val_str.split("\n")
                        val_str = max(lines, key=len)
                    max_len = max(max_len, len(val_str))
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def export_to_csv(self, report_data: dict, test_type: str = "all") -> str:
        """
        Exporta los resultados a formato CSV estándar con codificación UTF-8.
        """
        output = io.StringIO()
        # UTF-8 BOM para que Excel en español reconozca caracteres con tildes
        output.write("\ufeff")
        writer = csv.writer(output, delimiter=",", quoting=csv.QUOTE_MINIMAL)

        writer.writerow(["Tipo de Prueba", "N°", "Nombre Completo", "Cédula", "Edad", "Área / Puesto", "Resultado / Diagnóstico"])

        if test_type in ("all", "audiometria"):
            for i, r in enumerate(report_data.get("resultados_audiometria", []), start=1):
                writer.writerow([
                    "Audiometría",
                    i,
                    r.get("name", ""),
                    r.get("cedula", ""),
                    r.get("age", ""),
                    r.get("position", ""),
                    r.get("result", "")
                ])

        if test_type in ("all", "espirometria"):
            for i, r in enumerate(report_data.get("resultados_espirometria", []), start=1):
                writer.writerow([
                    "Espirometría",
                    i,
                    r.get("name", ""),
                    r.get("cedula", ""),
                    r.get("age", ""),
                    r.get("position", ""),
                    r.get("result", "")
                ])

        return output.getvalue()

    def export_full_backup_zip(self, data_root: Path) -> Tuple[Path, str]:
        """
        Crea una copia de seguridad integral (.caitbackup) de toda la base de datos,
        TODOS los borradores guardados, plantillas, archivos y PDFs adjuntos del sistema.
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"CAIT_Backup_Completo_{timestamp}.caitbackup"
        temp_dir = Path(tempfile.mkdtemp(prefix="cait_backup_"))
        backup_path = temp_dir / filename

        drafts_list = [f.name for f in (data_root / "reports").glob("*.json")] if (data_root / "reports").exists() else []
        databases_list = [f.name for f in (data_root / "databases").glob("*.json")] if (data_root / "databases").exists() else []
        attachments_count = len([f for f in (data_root / "attachments").rglob("*") if f.is_file()]) if (data_root / "attachments").exists() else 0

        with zipfile.ZipFile(backup_path, "w", zipfile.ZIP_DEFLATED) as zf:
            meta = {
                "backup_type": "FULL_SYSTEM_BACKUP",
                "version": self.VERSION,
                "created_at": datetime.now().isoformat(),
                "drafts_count": len(drafts_list),
                "drafts_list": drafts_list,
                "databases_list": databases_list,
                "attachments_count": attachments_count,
            }
            zf.writestr("backup_manifest.json", json.dumps(meta, indent=2))

            # Archivar todas las carpetas dentro de data_root
            for item in data_root.rglob("*"):
                if item.is_file():
                    rel_path = item.relative_to(data_root)
                    rel_str = str(rel_path).replace("\\", "/")
                    # En exports, incluir solo los PDFs generados, evitando archivos ZIP o caitbackup antiguos
                    if rel_str.startswith("exports"):
                        if item.suffix.lower() == ".pdf":
                            zf.write(item, arcname=f"data/{rel_path}")
                    else:
                        zf.write(item, arcname=f"data/{rel_path}")

        return backup_path, filename

    # =========================================================================
    # IMPORTACIÓN Y AUTO-REGISTRO
    # =========================================================================

    def import_report_package(
        self,
        file_bytes: bytes,
        filename: str,
        data_root: Path,
        persons_repo,
        evaluators_repo,
        counterparts_repo,
        normalize_func
    ) -> dict:
        """
        Importa un archivo (.cait, .json, o .caitpkg/.zip) y:
        1. Desempaqueta y normaliza los datos del reporte.
        2. Si es una copia de seguridad (.caitbackup o ZIP de sistema), restaura todos los borradores y bases de datos automáticamente.
        3. Restaura todos los adjuntos y PDFs que vengan en el paquete (físicos o embebidos en base64).
        4. Restaura todos los borradores incluidos a data_root/reports/.
        5. Registra el informe en data_root/reports/ (como actual y como borrador permanente).
        6. Registra/actualiza automáticamente cada paciente en el catálogo maestro (persons.json).
        7. Registra evaluadores y contrapartes nuevos si están presentes.
        """
        report_dict = {}
        associated_persons = []
        is_zip = False
        attachments_restored = 0
        drafts_restored = 0

        # Detectar si es archivo ZIP/caitpkg/caitbackup
        try:
            with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
                is_zip = True
                namelist = zf.namelist()
                
                # Auto-detección: Si es un backup del sistema completo, procesar como tal
                if "backup_manifest.json" in namelist or any(n.startswith("data/databases/") for n in namelist):
                    return self.restore_system_backup(file_bytes, data_root, mode="merge")

                # Buscar manifest o json principal
                target_json = None
                for n in namelist:
                    if n in ("manifest.json", "report.json", "current_report.json") or n.endswith(".cait") or (n.endswith(".json") and not n.startswith("drafts/")):
                        target_json = n
                        break
                
                if not target_json:
                    return {"status": "error", "message": "El archivo ZIP no contiene un reporte válido (.json / manifest.json)."}
                
                raw_json = zf.read(target_json).decode("utf-8")
                pkg_data = json.loads(raw_json)
                
                if isinstance(pkg_data, dict):
                    if "report" in pkg_data:
                        report_dict = pkg_data["report"]
                        associated_persons = pkg_data.get("associated_persons", [])
                    else:
                        report_dict = pkg_data
                        
                # 1. Extraer TODOS los adjuntos del ZIP a data_root/attachments/
                for name in namelist:
                    if name.startswith("attachments/") and not name.endswith("/"):
                        clean_sub = name[len("attachments/"):]
                        dest = data_root / "attachments" / clean_sub
                        dest.parent.mkdir(parents=True, exist_ok=True)
                        dest.write_bytes(zf.read(name))
                        attachments_restored += 1
                        
                # 2. Extraer TODOS los borradores incluidos en drafts/ a data_root/reports/
                for name in namelist:
                    if name.startswith("drafts/") and not name.endswith("/"):
                        draft_name = os.path.basename(name)
                        dest = data_root / "reports" / draft_name
                        dest.parent.mkdir(parents=True, exist_ok=True)
                        dest.write_bytes(zf.read(name))
                        drafts_restored += 1

                # 3. Extraer informes PDFs si vienen en exports/
                for name in namelist:
                    if name.startswith("exports/") and not name.endswith("/"):
                        clean_sub = name[len("exports/"):]
                        dest = data_root / "exports" / clean_sub
                        dest.parent.mkdir(parents=True, exist_ok=True)
                        dest.write_bytes(zf.read(name))

        except zipfile.BadZipFile:
            # Es un archivo plano JSON o .cait
            try:
                content = file_bytes.decode("utf-8-sig")
                pkg_data = json.loads(content)
                if isinstance(pkg_data, dict):
                    if "report" in pkg_data:
                        report_dict = pkg_data["report"]
                        associated_persons = pkg_data.get("associated_persons", [])
                    else:
                        report_dict = pkg_data
                        
                    # Extraer archivos/PDFs embebidos en base64 si existen
                    for finfo in pkg_data.get("embedded_files", []):
                        b64_data = finfo.get("b64")
                        fname = finfo.get("name")
                        category = finfo.get("category", "report_adjuntos")
                        if b64_data and fname:
                            dest = data_root / "attachments" / category / fname
                            dest.parent.mkdir(parents=True, exist_ok=True)
                            dest.write_bytes(base64.b64decode(b64_data))
                            attachments_restored += 1

                    # Extraer borradores incluidos si existen
                    for dinfo in pkg_data.get("saved_drafts", []):
                        dname = dinfo.get("name")
                        dcontent = dinfo.get("content")
                        if dname and dcontent:
                            dest = data_root / "reports" / dname
                            dest.parent.mkdir(parents=True, exist_ok=True)
                            dest.write_text(json.dumps(dcontent, ensure_ascii=False, indent=4), encoding="utf-8")
                            drafts_restored += 1
                else:
                    return {"status": "error", "message": "Estructura JSON inválida para importar reporte."}
            except Exception as e:
                return {"status": "error", "message": f"Error al decodificar archivo JSON: {e}"}

        # Normalizar el reporte
        norm_report = normalize_func(report_dict)
        
        # 1. Guardar como reporte actual
        current_path = data_root / "reports" / "current_report.json"
        current_path.parent.mkdir(parents=True, exist_ok=True)
        with open(current_path, "w", encoding="utf-8") as f:
            json.dump(norm_report, f, ensure_ascii=False, indent=4)

        # 2. Guardar también como borrador permanente con nombre de la empresa
        company = str(norm_report.get("company_name") or norm_report.get("company") or "Informe_Importado").strip()
        safe_comp = "".join(c for c in company if c.isalnum() or c in (" ", "_", "-")).strip().replace(" ", "_")
        eval_date = str(norm_report.get("evaluation_date") or "").replace("-", "")
        draft_filename = f"{safe_comp}_{eval_date or 'importado'}.json"
        draft_path = data_root / "reports" / draft_filename
        with open(draft_path, "w", encoding="utf-8") as f:
            json.dump(norm_report, f, ensure_ascii=False, indent=4)

        # 3. Auto-registro de personas en la base de datos (PersonsRepository)
        registered_count = 0
        
        # Primero registrar desde associated_persons si venían en el paquete
        for p in associated_persons:
            if isinstance(p, dict) and p.get("identification"):
                try:
                    persons_repo.upsert(p)
                    registered_count += 1
                except Exception:
                    pass

        # Luego registrar cada fila de audiometría y espirometría
        for item in norm_report.get("resultados_audiometria", []):
            if isinstance(item, dict) and item.get("cedula"):
                try:
                    persons_repo.upsert({
                        "identification": item["cedula"],
                        "name": item.get("name", ""),
                        "age": str(item.get("age", "")),
                        "position": item.get("position", ""),
                        "last_result_label": item.get("result", ""),
                        "last_test_type": "audiometria"
                    })
                    registered_count += 1
                except Exception:
                    pass

        for item in norm_report.get("resultados_espirometria", []):
            if isinstance(item, dict) and item.get("cedula"):
                try:
                    persons_repo.upsert({
                        "identification": item["cedula"],
                        "name": item.get("name", ""),
                        "age": str(item.get("age", "")),
                        "position": item.get("position", ""),
                        "last_result_label": item.get("result", ""),
                        "last_test_type": "espirometria"
                    })
                    registered_count += 1
                except Exception:
                    pass

        # 4. Auto-registro de evaluadores y contrapartes
        new_evaluators_count = 0
        new_counterparts_count = 0
        try:
            # Contraparte
            cp_name = norm_report.get("counterpart_name")
            cp_role = norm_report.get("counterpart_role")
            if cp_name:
                existing_cps = counterparts_repo.list_all()
                if not any(c.get("name", "").strip().lower() == cp_name.strip().lower() for c in existing_cps):
                    counterparts_repo.add_counterpart({"name": cp_name, "role": cp_role or "Contraparte Técnica"})
                    new_counterparts_count += 1

            # Perfil de evaluador si venía embebido
            ev_prof = norm_report.get("evaluator_profile")
            if isinstance(ev_prof, dict) and ev_prof.get("name"):
                existing_evs = evaluators_repo.list_all()
                if not any(e.get("name", "").strip().lower() == ev_prof["name"].strip().lower() for e in existing_evs):
                    evaluators_repo.add_evaluator(ev_prof)
                    new_evaluators_count += 1
        except Exception as e:
            print(f"Error registrando evaluadores/contrapartes importados: {e}")

        total_audio = len(norm_report.get("resultados_audiometria", []))
        total_espiro = len(norm_report.get("resultados_espirometria", []))

        return {
            "status": "ok",
            "message": f"¡Informe de '{company}' cargado y registrado exitosamente! ({attachments_restored} archivos/PDFs y {drafts_restored + 1} borradores guardados)",
            "details": {
                "draft_filename": draft_filename,
                "company": company,
                "total_audiometria": total_audio,
                "total_espirometria": total_espiro,
                "persons_registered": registered_count,
                "attachments_restored": attachments_restored,
                "drafts_restored": drafts_restored + 1,
                "new_evaluators": new_evaluators_count,
                "new_counterparts": new_counterparts_count,
            },
            "report": norm_report
        }

    def import_from_tabular_file(
        self,
        file_bytes: bytes,
        filename: str,
        target_test_type: str,
        persons_repo,
        default_result: str = ""
    ) -> dict:
        """
        Importa datos de pacientes desde Excel (.xlsx/.xls) o CSV (.csv).
        Detecta encabezados flexibles (Nombre, Cédula, Edad, Puesto, Resultado),
        valida y registra a cada persona en el catálogo maestro y devuelve
        la lista de filas listas para agregar a la tabla de resultados.
        """
        ext = Path(filename).suffix.lower()
        rows_data: List[Dict[str, Any]] = []

        if ext in (".xlsx", ".xls"):
            try:
                wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
                ws = wb.active
                
                # Leer todas las filas
                all_rows = list(ws.iter_rows(values_only=True))
                if not all_rows:
                    return {"status": "error", "message": "La hoja de Excel está vacía."}

                # Buscar fila de encabezado
                header_row_idx = 0
                headers = []
                for idx, r in enumerate(all_rows[:10]):
                    r_str = [str(c).lower() if c is not None else "" for c in r]
                    if any("nombre" in c or "name" in c or "cedula" in c or "cédula" in c for c in r_str):
                        header_row_idx = idx
                        headers = r_str
                        break
                
                if not headers:
                    headers = [str(c).lower() if c is not None else "" for c in all_rows[0]]
                    header_row_idx = 0

                # Mapear índices de columnas
                col_map = self._map_tabular_columns(headers)
                
                # Procesar filas de datos
                for r in all_rows[header_row_idx + 1:]:
                    if not any(r):
                        continue
                    entry = self._extract_row_data(r, col_map, target_test_type, default_result)
                    if entry and entry.get("name") and entry.get("cedula"):
                        rows_data.append(entry)

            except Exception as e:
                return {"status": "error", "message": f"Error leyendo archivo Excel: {e}"}

        elif ext == ".csv":
            try:
                text_content = file_bytes.decode("utf-8-sig", errors="replace")
                # Detectar delimitador (coma o punto y coma)
                first_line = text_content.splitlines()[0] if text_content.splitlines() else ""
                delimiter = ";" if first_line.count(";") > first_line.count(",") else ","
                
                reader = csv.reader(io.StringIO(text_content), delimiter=delimiter)
                all_rows = list(reader)
                if not all_rows:
                    return {"status": "error", "message": "El archivo CSV está vacio."}

                headers = [c.lower().strip() for c in all_rows[0]]
                col_map = self._map_tabular_columns(headers)

                for r in all_rows[1:]:
                    if not any(r):
                        continue
                    entry = self._extract_row_data(r, col_map, target_test_type, default_result)
                    if entry and entry.get("name") and entry.get("cedula"):
                        rows_data.append(entry)

            except Exception as e:
                return {"status": "error", "message": f"Error leyendo archivo CSV: {e}"}
        else:
            return {"status": "error", "message": f"Formato no compatible: {ext}. Use .xlsx, .xls o .csv"}

        if not rows_data:
            return {
                "status": "error",
                "message": "No se encontraron filas válidas con Nombre y Cédula en el archivo."
            }

        # Auto-registro en PersonsRepository
        registered_count = 0
        for item in rows_data:
            try:
                persons_repo.upsert({
                    "identification": item["cedula"],
                    "name": item["name"],
                    "age": str(item.get("age", "")),
                    "position": item.get("position", ""),
                    "last_result_label": item.get("result", ""),
                    "last_test_type": target_test_type
                })
                registered_count += 1
            except Exception:
                pass

        return {
            "status": "ok",
            "message": f"Se procesaron {len(rows_data)} pacientes y se registraron en la base de datos.",
            "total_extracted": len(rows_data),
            "persons_registered": registered_count,
            "entries": rows_data,
            "target_test_type": target_test_type
        }

    def restore_system_backup(self, file_bytes: bytes, data_root: Path, mode: str = "merge") -> dict:
        """
        Restaura o fusiona una copia de seguridad completa (.caitbackup/.zip) en el sistema,
        restaurando íntegramente todos los borradores, archivos y PDFs adjuntos y bases de datos.
        """
        try:
            with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
                namelist = zf.namelist()
                
                # Verificar si es backup de CAIT
                has_manifest = "backup_manifest.json" in namelist or any(n.startswith("data/") for n in namelist)
                if not has_manifest:
                    return {"status": "error", "message": "El archivo seleccionado no es un backup válido de CAIT."}

                restored_files = 0
                restored_drafts = 0
                restored_attachments = 0
                restored_databases = set()
                draft_names = []

                for item in namelist:
                    if item == "backup_manifest.json" or item.endswith("/"):
                        continue
                    
                    # Remover prefijo data/ si existe
                    clean_rel = item[5:] if item.startswith("data/") else item
                    clean_rel = clean_rel.replace("\\", "/")
                    target_path = data_root / clean_rel
                    
                    # Caso A: Borradores (reports/)
                    if clean_rel.startswith("reports/"):
                        target_path.parent.mkdir(parents=True, exist_ok=True)
                        target_path.write_bytes(zf.read(item))
                        restored_drafts += 1
                        draft_names.append(target_path.name)
                        restored_files += 1
                        continue

                    # Caso B: Archivos y PDFs adjuntos (attachments/)
                    if clean_rel.startswith("attachments/"):
                        target_path.parent.mkdir(parents=True, exist_ok=True)
                        target_path.write_bytes(zf.read(item))
                        restored_attachments += 1
                        restored_files += 1
                        continue

                    # Caso C: Informes PDF generados (exports/)
                    if clean_rel.startswith("exports/"):
                        target_path.parent.mkdir(parents=True, exist_ok=True)
                        target_path.write_bytes(zf.read(item))
                        restored_files += 1
                        continue

                    # Caso D: Bases de datos (databases/) con modo fusión
                    if clean_rel.startswith("databases/") and mode == "merge" and target_path.exists() and target_path.suffix == ".json":
                        try:
                            existing_data = json.loads(target_path.read_text(encoding="utf-8"))
                            backup_data = json.loads(zf.read(item).decode("utf-8"))
                            
                            if isinstance(existing_data, dict) and isinstance(backup_data, dict):
                                existing_data.update(backup_data)
                                target_path.write_text(json.dumps(existing_data, ensure_ascii=False, indent=2), encoding="utf-8")
                            elif isinstance(existing_data, list) and isinstance(backup_data, list):
                                seen_keys = set()
                                for x in existing_data:
                                    if isinstance(x, dict):
                                        k = x.get("id") or x.get("identification") or x.get("name")
                                        if k: seen_keys.add(str(k).strip().upper())
                                for b_item in backup_data:
                                    if isinstance(b_item, dict):
                                        k = b_item.get("id") or b_item.get("identification") or b_item.get("name")
                                        if not k or str(k).strip().upper() not in seen_keys:
                                            existing_data.append(b_item)
                                            if k: seen_keys.add(str(k).strip().upper())
                                target_path.write_text(json.dumps(existing_data, ensure_ascii=False, indent=2), encoding="utf-8")
                            else:
                                target_path.write_bytes(zf.read(item))
                            restored_databases.add(target_path.name)
                        except Exception as e:
                            print(f"Error fusionando base de datos {target_path}: {e}")
                            target_path.write_bytes(zf.read(item))
                    else:
                        target_path.parent.mkdir(parents=True, exist_ok=True)
                        target_path.write_bytes(zf.read(item))
                        if clean_rel.startswith("databases/"):
                            restored_databases.add(target_path.name)
                        
                    restored_files += 1

                return {
                    "status": "ok",
                    "message": f"Copia de seguridad restaurada exitosamente ({restored_drafts} borradores, {restored_attachments} archivos/PDFs adjuntos, {len(restored_databases)} bases de datos actualizadas).",
                    "details": {
                        "drafts_restored": restored_drafts,
                        "attachments_restored": restored_attachments,
                        "databases_merged": len(restored_databases),
                        "total_files": restored_files,
                        "draft_names": draft_names
                    }
                }

        except Exception as e:
            return {"status": "error", "message": f"Error al procesar la copia de seguridad: {e}"}

    # =========================================================================
    # HELPERS INTERNOS DE MAPEO
    # =========================================================================

    def _map_tabular_columns(self, headers: List[str]) -> Dict[str, int]:
        """Detecta las posiciones de columnas basándose en variantes comunes."""
        col_map = {}
        for idx, h in enumerate(headers):
            h_clean = h.strip().lower()
            
            # Nombre
            if any(k in h_clean for k in ("nombre", "name", "trabajador", "colaborador", "empleado", "paciente")):
                if "name" not in col_map:
                    col_map["name"] = idx
            # Cédula
            elif any(k in h_clean for k in ("cedula", "cédula", "identificacion", "identificación", "id", "documento")):
                if "cedula" not in col_map:
                    col_map["cedula"] = idx
            # Edad
            elif any(k in h_clean for k in ("edad", "age", "años")):
                if "age" not in col_map:
                    col_map["age"] = idx
            # Puesto / Cargo / Área
            elif any(k in h_clean for k in ("puesto", "cargo", "area", "área", "departamento", "posicion", "position", "ocupacion")):
                if "position" not in col_map:
                    col_map["position"] = idx
            # Resultado / Diagnóstico
            elif any(k in h_clean for k in ("resultado", "diagnostico", "diagnóstico", "result", "conclusion")):
                if "result" not in col_map:
                    col_map["result"] = idx

        return col_map

    def _extract_row_data(
        self,
        row: tuple | list,
        col_map: Dict[str, int],
        test_type: str,
        default_result: str
    ) -> Optional[Dict[str, str]]:
        """Extrae y normaliza un registro individual a partir de la fila y el mapa de columnas."""
        def get_val(key):
            if key in col_map and col_map[key] < len(row):
                v = row[col_map[key]]
                return str(v).strip() if v is not None else ""
            return ""

        name = get_val("name")
        cedula = get_val("cedula").upper()
        age = get_val("age")
        position = get_val("position")
        result = get_val("result")

        # Limpiar edad si viene como float (ej. 35.0 en Excel)
        if age.endswith(".0"):
            age = age[:-2]

        if not name or not cedula:
            return None

        # Si no hay resultado en la columna, asignar el por defecto según el tipo
        if not result:
            if default_result:
                result = default_result
            elif "espiro" in test_type.lower():
                result = "Espirometría normal"
            else:
                result = "Normal bilateral"

        return {
            "name": name,
            "cedula": cedula,
            "age": age,
            "position": position or "General",
            "result": result
        }
